import openpyxl, glob, os, re

GLOTTOLOG_NEW = {
    'aben1249': ('VERIFICAR',               'VERIFICAR'),           # Nitsch 2009
    'achi1257': ('Acehnese',                'Austronesian'),
    'adny1235': ('Adnyamathanha',           'Pama-Nyungan'),
    'aghu1254': ('Aghu',                    'Trans-New Guinea'),
    'agus1235': ('Agusan Manobo',           'Austronesian'),
    'ahin1234': ('VERIFICAR',               'VERIFICAR'),           # Santiago 2013
    'aime1238': ('VERIFICAR',               'VERIFICAR'),           # Aiton 2016
    'akun1241': ('VERIFICAR',               'VERIFICAR'),           # Aragon 2014
    'alba1269': ('Alaba',                   'Afro-Asiatic'),        # Fincke 2002
    'aleu1260': ('Aleut',                   'Eskimo-Aleut'),
    'alya1239': ('Alyawarra',               'Arandic'),
    'amam1246': ('VERIFICAR',               'VERIFICAR'),           # Aki & Pennington 2013
    'amdo1237': ('Amdo Tibetan',            'Trans-Himalayan'),
    'amis1246': ('Amis',                    'Austronesian'),
    'ango1257': ('VERIFICAR',               'VERIFICAR'),           # Casilimas Rojas 1995
    'anut1237': ('Anutan',                  'Austronesian'),
    'apat1240': ('Apatani',                 'Trans-Himalayan'),
    'apin1244': ('Apinajé',                 'Macro-Jê'),
    'arab1267': ('Arabana',                 'Pama-Nyungan'),
    'arta1239': ('Arta',                    'Austronesian'),
    'athp1241': ('Athpare',                 'Trans-Himalayan'),
    'auye1238': ('Auye',                    'Trans-New Guinea'),
    'awab1243': ('Awabakal',                'Pama-Nyungan'),
    'awar1248': ('Avar',                    'Nakh-Daghestanian'),
    'ayiw1239': ('Äiwoo',                   'Austronesian'),
    'bahi1252': ('Bahing',                  'Trans-Himalayan'),
    'bala1311': ('Balangao',                'Austronesian'),
    'bala1316': ('VERIFICAR',               'Austronesian'),        # Ozanne-Rivierre 1998, lengua Kanak
    'bamb1270': ('VERIFICAR',               'VERIFICAR'),           # Campbell 1989
    'bant1281': ('Bantawa',                 'Trans-Himalayan'),
    'bara1357': ('VERIFICAR',               'VERIFICAR'),           # Dhakal 2014
    'basq1248': ('Basque',                  'Language isolate'),
    'bawm1236': ('Bawm Chin',               'Trans-Himalayan'),
    'bayb1234': ('VERIFICAR',               'VERIFICAR'),           # Rubino 2005
    'belh1239': ('Belhare',                 'Trans-Himalayan'),
    'bert1248': ('Berta',                   'Nilo-Saharan'),
    'bier1244': ('Bierebo',                 'Austronesian'),
    'bilb1241': ('VERIFICAR',               'VERIFICAR'),           # Dempwolff 1909
    'bina1277': ('Binahari',                'Trans-New Guinea'),
    'binu1244': ('Binukid',                 'Austronesian'),
    'biri1256': ('Biri',                    'Pama-Nyungan'),
    'boun1245': ('VERIFICAR',               'VERIFICAR'),           # Merlan & Rumsey 1991
    'bugi1244': ('Bugis',                   'Austronesian'),
    'bujh1238': ('Bujheli',                 'Trans-Himalayan'),
    'bund1253': ('Bundeli',                 'Indo-European'),
    'caml1239': ('Camling',                 'Trans-Himalayan'),
    'casi1235': ('Casiguran Dumagat Agta',  'Austronesian'),
    'cent2127': ('Central Alaskan Yupik',   'Eskimo-Aleut'),
    'chan1313': ('Chang Naga',              'Trans-Himalayan'),
    'chep1245': ('Chepang',                 'Trans-Himalayan'),
    'choc1275': ('Chocangacakha',           'Trans-Himalayan'),
    'chug1252': ('Duhumbi',                 'Trans-Himalayan'),
    'chuk1273': ('Chukchi',                 'Chukotko-Kamchatkan'),
    'colu1241': ('Nxaamxcin',               'Salishan'),
    'coos1249': ('Hanis Coos',              'Coosan'),
    'cuti1242': ('VERIFICAR',               'VERIFICAR'),           # Dos Anjos 2011
    'daai1236': ('Daai Chin',               'Trans-Himalayan'),
    'dara1250': ('Darai',                   'Indo-European'),
    'darl1243': ('Paakantyi',               'Pama-Nyungan'),        # Hercus 1982
    'darm1243': ('Darma',                   'Trans-Himalayan'),
    'gahr1239': ('Bunan',                   'Trans-Himalayan'),
    'gana1278': ('Ganalbingu',              'Pama-Nyungan'),
    'haka1240': ('Haka Chin',               'Trans-Himalayan'),
    'ilon1239': ('Ilongot',                 'Austronesian'),
    'jeru1240': ('Jerung',                  'Trans-Himalayan'),
    'kalk1246': ('Kalkatungu',              'Pama-Nyungan'),
    'kamu1260': ('Kamula',                  'Trans-New Guinea'),
    'kara1476': ('VERIFICAR',               'VERIFICAR'),           # McKelson 2004
    'kari1254': ('Dzubukua',                'Karirí'),
    'kari1311': ('Karitiana',               'Tupian'),
    'noct1238': ('Nocte Naga',              'Trans-Himalayan'),
    'nort2885': ('Northern Subanen',        'Austronesian'),
}

FILENAME_TO_GLOTTOCODE = {
    'Shiriana_language_ergativity': 'nina1238',
}

def extract_glottocode(stem):
    if stem in FILENAME_TO_GLOTTOCODE:
        return FILENAME_TO_GLOTTOCODE[stem]
    m = re.search(r'[a-z]{4}[0-9]{4}', stem)
    return m.group(0) if m else None

def add_columns(filepath, glottocode, lang, family):
    wb = openpyxl.load_workbook(filepath)
    for sname in wb.sheetnames:
        ws = wb[sname]
        ws.insert_cols(1, 3)
        ws.cell(row=1, column=1, value='Glottocode')
        ws.cell(row=1, column=2, value='Language')
        ws.cell(row=1, column=3, value='Family')
        for r in range(2, ws.max_row + 1):
            if any(ws.cell(row=r, column=c).value is not None for c in range(4, ws.max_column+1)):
                ws.cell(row=r, column=1, value=glottocode)
                ws.cell(row=r, column=2, value=lang)
                ws.cell(row=r, column=3, value=family)
    wb.save(filepath)

folder = '/sessions/compassionate-lucid-gauss/mnt/Ergon/fichas_ergon/'
# Solo archivos SIN columnas aún
files = [f for f in sorted(glob.glob(folder + '*.xlsx'))
         if openpyxl.load_workbook(f, read_only=True).active.cell(1,1).value != 'Glottocode']

ok, verif, err = 0, 0, 0
for f in files:
    stem = os.path.basename(f).replace('.xlsx','')
    gc = extract_glottocode(stem)
    if not gc:
        print(f'[NO CODE] {stem}'); err += 1; continue
    if gc not in GLOTTOLOG_NEW:
        lang, fam = f'VERIFICAR ({gc})', 'VERIFICAR'
        print(f'[NUEVO?] {stem} → {gc}'); verif += 1
    else:
        lang, fam = GLOTTOLOG_NEW[gc]
    if 'VERIFICAR' in lang: verif += 1
    try:
        add_columns(f, gc, lang, fam)
        ok += 1
        print(f'[OK] {stem} → {gc} | {lang} | {fam}')
    except Exception as e:
        print(f'[ERROR] {stem}: {e}'); err += 1

print(f'\n=== RESUMEN LOTE 2 === OK:{ok} | VERIFICAR:{verif} | ERROR:{err}')
