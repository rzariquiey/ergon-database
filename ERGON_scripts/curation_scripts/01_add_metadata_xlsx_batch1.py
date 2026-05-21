import openpyxl
import glob
import os
import re

# ─────────────────────────────────────────────────────
# Tabla de metadatos Glottolog
# Basada en fuentes bibliográficas de cada ficha y
# conocimiento del catálogo Glottolog 4.x
# Entradas marcadas con "VERIFICAR" requieren revisión manual
# ─────────────────────────────────────────────────────
GLOTTOLOG = {
    # glottocode:  (nombre_lengua,             familia)
    'nucl1302': ('Nuclear Malagasy',            'Austronesian'),         # Plantilla
    'nina1238': ('Ninam',                       'Yanomaman'),            # Shiriana_language_ergativity
    'byan1241': ('Byansi',                      'Trans-Himalayan'),
    'caca1251': ('Kakataibo',                   'Panoan'),
    'cavi1250': ('Cavineña',                    'Tacanan'),
    'cemu1238': ('Cemuhî',                      'Austronesian'),
    'chac1251': ('Chacobo',                     'Panoan'),
    'coch1272': ('Cochimí',                     'Cochimí-Yuman'),
    'coeu1236': ("Coeur d'Alene",               'Salishan'),
    'dech1234': ('VERIFICAR (ficha dice dadi1250 Dadibi)', 'VERIFICAR'),
    'dhar1247': ('Thargari',                    'Pama-Nyungan'),
    'dhar1248': ('VERIFICAR',                   'VERIFICAR'),           # Terrill 2002
    'dier1241': ('Dieri',                       'Pama-Nyungan'),
    'diga1241': ('Idu Mishmi',                  'Trans-Himalayan'),
    'djam1255': ('Jaminjung',                   'Mirndi'),
    'djap1238': ('Djapu',                       'Pama-Nyungan'),
    'djin1251': ('Jingulu',                     'Mirndi'),
    'doma1260': ('Domaaki',                     'Indo-European'),
    'drun1238': ('Drung',                       'Trans-Himalayan'),
    'dumi1241': ('Dumi',                        'Trans-Himalayan'),
    'dura1244': ('Dura',                        'Trans-Himalayan'),
    'duun1241': ('Duungidjawu',                 'Pama-Nyungan'),
    'dyan1250': ('Dyangadi',                    'Pama-Nyungan'),
    'dyir1250': ('Dyirbal',                     'Pama-Nyungan'),
    'dzon1239': ('Dzongkha',                    'Trans-Himalayan'),
    'east2343': ('Yohlmo',                      'Trans-Himalayan'),
    'east2379': ('Mparntwe Arrernte',           'Arandic'),
    'east2447': ('East Futunan',                'Austronesian'),
    'east2516': ('East Kewa',                   'Trans-New Guinea'),
    'east2545': ('Eastern Pomo',                'Pomoan'),
    'east2694': ('Eastern Subanun',             'Austronesian'),
    'east2773': ('Dolakha Newar',               'Trans-Himalayan'),
    'east2851': ('Geshiza',                     'Trans-Himalayan'),
    'emai1242': ('Emai',                        'Niger-Congo'),
    'ersu1241': ('Ersu',                        'Trans-Himalayan'),
    'fala1243': ('Falam Chin',                  'Trans-Himalayan'),
    'fass1245': ('Fassano',                     'Indo-European'),
    'fasu1242': ('Fasu',                        'Trans-New Guinea'),
    'fran1266': ('VERIFICAR',                   'VERIFICAR'),           # Engel & Bartholomew 1987
    'fuln1247': ('Fulniô',                      'Fulniô'),
    'futu1245': ('Futuna-Aniwa',                'Austronesian'),
    'guan1266': ('Guanano',                     'Tucanoan'),
    'gugu1255': ('Guugu Yimithirr',             'Pama-Nyungan'),
    'guiq1238': ('Guiqiong',                    'Trans-Himalayan'),
    'guny1241': ('Gunya',                       'Pama-Nyungan'),
    'gure1255': ('Gure-Kahugu',                 'Niger-Congo'),
    'guru1261': ('Gurung',                      'Trans-Himalayan'),
    'hida1246': ('Hidatsa',                     'Siouan'),
    'hind1269': ('Hindi',                       'Indo-European'),
    'hual1240': ('Hualapai',                    'Cochimí-Yuman'),
    'huas1242': ('Huastec',                     'Mayan'),
    'hula1239': ('Hula',                        'Austronesian'),
    'huli1244': ('Huli',                        'Trans-New Guinea'),
    'hunz1247': ('Hunzib',                      'Nakh-Daghestanian'),
    'hurr1240': ('Hurrian',                     'Hurro-Urartian'),
    'iaai1238': ('Iaai',                        'Austronesian'),
    'ibal1244': ('Ibaloi',                      'Austronesian'),
    'iban1267': ('Iban',                        'Austronesian'),
    'isco1239': ('Iskonawa',                    'Panoan'),
    'isin1239': ('Isinai',                      'Austronesian'),
    'iwam1256': ('Iwam',                        'Sepik'),
    'japh1234': ('Japhug',                      'Trans-Himalayan'),
    'jaru1254': ('Jaru',                        'Pama-Nyungan'),
    'kaba1278': ('Kabardian',                   'Northwest Caucasian'),
    'kaga1256': ('Kagayanen',                   'Austronesian'),
    'kaik1246': ('Kaike',                       'Trans-Himalayan'),
    'kaki1249': ('Kaki Ae',                     'Trans-New Guinea'),
    'kala1399': ('Kalasha',                     'Indo-European'),
    'kank1243': ('Kankanaey',                   'Austronesian'),
    'kara1487': ('Karao',                       'Austronesian'),
    'kara1500': ('Karajá',                      'Macro-Jê'),
    'katu1276': ('Katu',                        'Austroasiatic'),
    'mati1255': ('Matis',                       'Panoan'),
    'mats1244': ('Matsés',                      'Panoan'),
    'nhan1238': ('Nhanda',                      'Pama-Nyungan'),
    'ning1281': ('Anong',                       'Trans-Himalayan'),
    'niue1239': ('Niuean',                      'Austronesian'),
    'nort2722': ('Northern Qiang',              'Trans-Himalayan'),
    'nort2745': ('Tundra Yukaghir',             'Yukaghir'),
    'panj1256': ('Punjabi',                     'Indo-European'),
    'patt1248': ('Pattani',                     'Indo-European'),
    'pemo1248': ('Pemón',                       'Cariban'),
    'pend1242': ('Pendau',                      'Austronesian'),
    'ship1254': ('Shipibo-Konibo',              'Panoan'),
    'yami1258': ('Yami',                        'Austronesian'),
}

# Archivos con glottocode no estándar en el nombre
FILENAME_TO_GLOTTOCODE = {
    'Shiriana_language_ergativity': 'nina1238',
}

def extract_glottocode(filename_stem):
    """Extrae el glottocode (4letras+4digitos) del nombre del archivo."""
    if filename_stem in FILENAME_TO_GLOTTOCODE:
        return FILENAME_TO_GLOTTOCODE[filename_stem]
    match = re.search(r'[a-z]{4}[0-9]{4}', filename_stem)
    if match:
        return match.group(0)
    return None

def add_columns_to_xlsx(filepath, glottocode, lang_name, family):
    """Inserta 3 columnas al inicio: Glottocode, Language, Family."""
    wb = openpyxl.load_workbook(filepath)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # Insertar 3 columnas al inicio (desplaza el contenido existente)
        ws.insert_cols(1, 3)
        # Cabecera en fila 1
        ws.cell(row=1, column=1, value='Glottocode')
        ws.cell(row=1, column=2, value='Language')
        ws.cell(row=1, column=3, value='Family')
        # Rellenar cada fila de datos con los metadatos
        for row_idx in range(2, ws.max_row + 1):
            # Solo rellenar si hay algún valor en la fila (no filas vacías)
            row_has_data = any(ws.cell(row=row_idx, column=c).value is not None
                               for c in range(4, ws.max_column + 1))
            if row_has_data:
                ws.cell(row=row_idx, column=1, value=glottocode)
                ws.cell(row=row_idx, column=2, value=lang_name)
                ws.cell(row=row_idx, column=3, value=family)
    wb.save(filepath)
    return True

# ─────────────────────────────────────────────────────
# Procesamiento de todos los xlsx
# ─────────────────────────────────────────────────────
folder = '/sessions/compassionate-lucid-gauss/mnt/Ergon/fichas_ergon/'
files = sorted(glob.glob(folder + '*.xlsx'))

ok_count = 0
warn_count = 0
error_count = 0
missing_code = []
verify_list = []

for filepath in files:
    stem = os.path.basename(filepath).replace('.xlsx', '')
    glottocode = extract_glottocode(stem)

    if glottocode is None:
        print(f'[NO CODE] {stem}')
        missing_code.append(stem)
        continue

    if glottocode not in GLOTTOLOG:
        print(f'[NOT IN TABLE] {stem} → {glottocode}')
        lang_name = f'VERIFICAR ({glottocode})'
        family = 'VERIFICAR'
        warn_count += 1
    else:
        lang_name, family = GLOTTOLOG[glottocode]
        if 'VERIFICAR' in lang_name or 'VERIFICAR' in family:
            verify_list.append((stem, glottocode))

    try:
        add_columns_to_xlsx(filepath, glottocode, lang_name, family)
        ok_count += 1
        print(f'[OK] {stem} → {glottocode} | {lang_name} | {family}')
    except Exception as e:
        print(f'[ERROR] {stem}: {e}')
        error_count += 1

print()
print(f'=== RESUMEN ===')
print(f'Procesados OK:    {ok_count}')
print(f'Avisos (en tabla pero VERIFICAR): {len(verify_list)}')
print(f'Fuera de tabla:   {warn_count}')
print(f'Sin glottocode:   {len(missing_code)}')
print(f'Errores:          {error_count}')
if verify_list:
    print(f'Entradas a verificar: {[v[0] for v in verify_list]}')
if missing_code:
    print(f'Sin código detectado: {missing_code}')
